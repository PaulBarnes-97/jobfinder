from openpyxl import Workbook
from openpyxl import load_workbook
import test
import os
from datetime import date
import json




def heading(cs):
    headings = ("Job Title", "Company", "City", "State", "Benefit")
    for i in range(0, len(headings)):
        cs.cell(row=1, column=i + 1).value = headings[i]



def validate(fp, t):
    if os.path.isfile(fp):
        cwb = load_workbook(fp, read_only=False)
        print(fp,' exists')
        if str(t) not in cwb.sheetnames:
            print('sheet does not exist, creating')
            cs = cwb.create_sheet(str(t))
            heading(cs)
            cwb.save(fp)
            print("Done, saved")
        else:
            print('sheet exists')
    else:
        print(fp, 'does not exist')
    print('Entry: ', str(t),"/Validated", "\n", "Searching for Entries:")

def filler(fp,t):

    cwb = load_workbook(fp, read_only=False)
    cws = cwb[str(t)]
    fpj = os.getcwd() + "/json dump/" + str(t) +".json"
    jso = json.loads(open(fpj).read())
    val = jso["job"]
    for i in range(0,len(val)):
        #"Job Title", "Company", "City", "State", "Benefit"
        r = i+2
        cws.cell(column=1, row=r).value='=HYPERLINK("{}","{}")'.format((val[i]["Link"]),(val[i]["Job_Title"]))
        cws.cell(column=2, row=r).value='=HYPERLINK("{}","{}")'.format((val[i]["LinkedIn"]), (val[i]["Company"]))
        cws.cell(column=3, row=r, value=val[i]["City"])
        cws.cell(column=4, row=r, value=val[i]["State"])
        cws.cell(column=5, row=r, value=val[i]["Benefits"])

        print(i)

    print('saving: ')
    cwb.save("C:\\Users\\Me not you\\Documents\\jobs.xlsx")
def main():
    file_path = "C:\\Users\\Me not you\\Documents\\jobs.xlsx"
    today = date.today()
    #test.job_search(today)
    validate(file_path, today)
    filler(file_path,today)

main()

