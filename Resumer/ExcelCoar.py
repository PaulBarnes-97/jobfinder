from openpyxl import Workbook
from openpyxl import load_workbook
import test
import time
import os
from datetime import date

file_path = "C:\\Users\\Me not you\\Documents\\jobs.xlsx"
today = date.today()
t = today


def heading(cs):
    headings = ("Visited","Time" ,"Job Title", "Company", "LinkedIn", "City", "State", "Benefit", "Link","ID", "Reply")
    for i in range(0, len(headings)):
        cs.cell(row=1, column=i + 1).value = headings[i]



def validate(fp, t):
    if os.path.isfile(fp):
        cwb = load_workbook(fp, read_only=False)
        print(fp,' exists')
        if str(t) not in cwb.sheetnames:
            print('sheet does not exist, creating')
            cs = cwb.create_sheet(str(t))
            heading(cs)

            cwb.save(fp)
            print("Done, saved")
        else:
            print('sheet exists')
    else:
        print(fp, 'does not exist')
    print('Entry: ', str(t),"/Validated", "\n", "Searching for Entries:")

def fill(j):

    cwb = load_workbook(file_path, read_only=False)
    cws = cwb[str(today)]
    for i in range(0,len(j)):
        "Visited", "Time", "Job Title", "Company", "LinkedIn", "City", "State", "Benefit", "Link", "ID", "Reply"
        c=0
        r = cws.max_row+1
        for k in range(3,10):
            cws.cell(column=k, row=r, value=str(j[i][c]))
            print(str(j[i][c]))
            c+=1
    cwb.save("C:\\Users\\Me not you\\Documents\\jobs.xlsx")
def main():
    validate(file_path, today)
    fill(test.job_search())
main()

